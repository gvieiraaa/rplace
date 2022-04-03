from PIL import Image
import openpyxl as xl
from openpyxl.styles import PatternFill

def rgb2hex(r, g, b):
    return f'{r:02X}{g:02X}{b:02X}'

im = Image.open('br_area.png')

wb = xl.Workbook()
ws = wb.active
ws1 = wb.create_sheet("Mysheet")


for x in range(im.width):
    for y in range(im.height):
        cell = ws.cell(row=y+1, column=x+1)
        cell.fill = PatternFill("solid", fgColor=rgb2hex(*im.getpixel((x,y))))

wb.save('test_export.xlsx')
